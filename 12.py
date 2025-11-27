import os
import sys
import json
import argparse
import subprocess
from collections import defaultdict
from datetime import datetime, date, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil

# pot do mape za poročila
REPORT_DIR = "poročila"
ARCHIVE_DIR = os.path.join(REPORT_DIR, "stara poročila in koledarji")

os.makedirs(REPORT_DIR, exist_ok=True)
os.makedirs(ARCHIVE_DIR, exist_ok=True)


def try_open_file(path: str, label: str):
    """Poskusi odpreti datoteko na več platformah brez izjem v primeru neuspeha."""
    try:
        if hasattr(os, "startfile"):
            os.startfile(path)
            return

        # Linux / macOS
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.run([opener, path], check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as e:  # pylint: disable=broad-except
        print(f"{label} ni bilo mogoče samodejno odpreti:", e)

def archive_old_files(prefix: str, keep_path: str):
    """
    Premakne vse datoteke v REPORT_DIR, ki se začnejo z `prefix`,
    in NISO nova datoteka `keep_path`, v ARCHIVE_DIR.
    """
    keep_abs = os.path.abspath(keep_path)

    for fname in os.listdir(REPORT_DIR):
        if not fname.lower().endswith(".xlsx"):
            continue
        if not fname.startswith(prefix):
            continue

        old_path = os.path.join(REPORT_DIR, fname)

        if os.path.abspath(old_path) == keep_abs:
            continue  # to je nova datoteka, ostane

        dest = os.path.join(ARCHIVE_DIR, fname)

        if os.path.exists(dest):
            base, ext = os.path.splitext(fname)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = os.path.join(ARCHIVE_DIR, f"{base}_{ts}{ext}")

        try:
            shutil.move(old_path, dest)
        except PermissionError:
            print(f"Datoteke {old_path} ni mogoče premakniti (verjetno odprta v Excelu).")

import re

def get_next_version_path(base_prefix: str) -> str:
    """
    Vrne pot do naslednje verzije datoteke v REPORT_DIR, ob upoštevanju
    obstoječih verzij v REPORT_DIR in ARCHIVE_DIR.
    Imena v obliki base_prefix_vN.xlsx (N = 1,2,3,...).
    """
    pattern = re.compile(re.escape(base_prefix) + r"_v(\d+)\.xlsx$", re.IGNORECASE)
    max_v = 0

    # preglej trenutna poročila
    for fname in os.listdir(REPORT_DIR):
        if not fname.lower().endswith(".xlsx"):
            continue
        if not fname.startswith(base_prefix):
            continue
        m = pattern.match(fname)
        if m:
            v = int(m.group(1))
            if v > max_v:
                max_v = v

    # preglej arhiv
    for fname in os.listdir(ARCHIVE_DIR):
        if not fname.lower().endswith(".xlsx"):
            continue
        if not fname.startswith(base_prefix):
            continue
        m = pattern.match(fname)
        if m:
            v = int(m.group(1))
            if v > max_v:
                max_v = v

    next_v = max_v + 1
    fname = f"{base_prefix}_v{next_v}.xlsx"
    return os.path.join(REPORT_DIR, fname)


# =============================
# ARGUMENTI / CONFIG
# =============================

def load_config(path: str):
    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    # periods
    periods = {}
    for rok, rng in cfg["periods"].items():
        start = datetime.strptime(rng["start"], "%Y-%m-%d").date()
        end = datetime.strptime(rng["end"], "%Y-%m-%d").date()
        periods[rok] = (start, end)

    # prazniki
    prazniki = set()
    for s in cfg.get("prazniki", []):
        prazniki.add(datetime.strptime(s, "%Y-%m-%d").date())

    # konfliktne skupine
    konfliktne = cfg.get("konfliktne_skupine", {})

    return periods, prazniki, konfliktne


def parse_args():
    p = argparse.ArgumentParser(description="Analiza izpitnih rokov FI.")
    p.add_argument(
        "--input",
        default="izpitni_roki_25_26_prilagojen_v5.xlsx",
        help="Vhodna Excel datoteka z roki."
    )
    p.add_argument(
        "--sheet",
        default="Profesorji_pripravljeno",
        help="Ime lista z roki."
    )
    p.add_argument(
        "--config",
        default="config_roki.json",
        help="Config JSON za obdobja, praznike, konfliktne skupine."
    )
    p.add_argument(
        "--out-porocilo",
        default=None,
        help="Ime poročila (xlsx). Če ni podano, se generira avtomatsko."
    )
    p.add_argument(
        "--out-koledar",
        default=None,
        help="Ime koledarja (xlsx). Če ni podano, se generira avtomatsko."
    )
    p.add_argument(
        "--rok-filter",
        choices=["zimsko", "poletno", "jesensko"],
        help="Obdelaj samo izbran rok."
    )
    p.add_argument(
        "--prof-filter",
        help="Obdelaj samo izbranega profesorja (substring match)."
    )
    p.add_argument(
        "--gui",
        action="store_true",
        help="(rezervirano) – preprost GUI, če boš želel v ločenem skriptu."
    )
    return p.parse_args()


args = parse_args()
PERIODS, PRAZNIKI, KONFLIKTNE_SKUPINE_PO_ROKU = load_config(args.config)

INPUT_FILE = args.input
SHEET_NAME = args.sheet

ZIMSKO_OD, ZIMSKO_DO = PERIODS["zimsko"]
POLETNO_OD, POLETNO_DO = PERIODS["poletno"]
JESENSKO_OD, JESENSKO_DO = PERIODS["jesensko"]

# =============================
# POMOŽNE FUNKCIJE
# =============================

def get_rok_range(rok: str):
    if rok not in PERIODS:
        return None, None
    return PERIODS[rok]


def get_rok(d: date) -> str | None:
    if d is None:
        return None
    for rok, (start, end) in PERIODS.items():
        if start <= d <= end:
            return rok
    return None


def describe_out_of_range(d: date) -> str:
    """Opis, katero mejo roka datum krši (grobo, za info)."""
    if d < ZIMSKO_OD:
        return f"pred ZIMSKO_OD ({ZIMSKO_OD})"
    if ZIMSKO_DO < d < POLETNO_OD:
        return f"po ZIMSKO_DO ({ZIMSKO_DO})"
    if POLETNO_DO < d < JESENSKO_OD:
        return f"po POLETNO_DO ({POLETNO_DO})"
    if d > JESENSKO_DO:
        return f"po JESENSKO_DO ({JESENSKO_DO})"
    return "znotraj katerega od obdobij (logična napaka)"


def normalize_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%d. %m. %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def normalize_predmet_name(name: str) -> str:
    if pd.isna(name):
        return ""
    s = str(name)
    s = s.replace("\xa0", " ")
    s = " ".join(s.split())
    return s.strip()


def expand_predmeti_row(row):
    """
    'Predmeti' = 'A. B. C' -> tri vrstice: A, B, C (isti datum, prof, ura ...).
    """
    predmeti_raw = row["Predmeti"]
    if pd.isna(predmeti_raw):
        return []

    parts = [p.strip() for p in str(predmeti_raw).split(".") if p.strip()]
    expanded = []
    for predmet in parts:
        new_row = row.copy()
        new_row["Predmet"] = predmet
        expanded.append(new_row)
    return expanded


def is_weekend(d: date) -> bool:
    if d is None:
        return False
    return d.weekday() >= 5  # 5 = sobota, 6 = nedelja


def predmet_groups_for_rok(rok: str, predmet: str):
    """Vrne seznam skupin (list[list[str]]) v katerih je ta predmet za dani rok."""
    if rok not in KONFLIKTNE_SKUPINE_PO_ROKU:
        return []
    groups = []
    for grp in KONFLIKTNE_SKUPINE_PO_ROKU[rok]:
        if predmet in grp:
            groups.append(grp)
    return groups


def build_proposal_context(df_all: pd.DataFrame):
    """Pripravi predizračune za generiranje predlogov datumov."""
    df_with_dates = df_all[df_all["Datum"].notna() & df_all["Rok"].notna()]

    dates_by_predmet_rok = defaultdict(list)
    rows_by_rok_date: dict[str, dict[date, list[dict]]] = defaultdict(
        lambda: defaultdict(list)
    )

    for idx, row in df_with_dates.iterrows():
        key = (row["Predmet"], row["Rok"])
        dates_by_predmet_rok[key].append((idx, row["Datum"]))

        rows_by_rok_date[row["Rok"]][row["Datum"]].append(
            {
                "Predmet": row["Predmet"],
                "Prof": row["Prof"],
                "Ura": row["Ura"],
                "Predavalnica": row["Predavalnica"],
            }
        )

    return {
        "dates_by_predmet_rok": dates_by_predmet_rok,
        "rows_by_rok_date": rows_by_rok_date,
    }


def propose_dates_for_row(row, context: dict, top_n: int = 5):
    """
    Za eno problematično vrstico predlaga do top_n novih datumov.
    Upošteva:
    - isti rok,
    - brez vikendov,
    - brez praznikov,
    - 10-dnevno pravilo znotraj istega roka (naprej + nazaj),
    - brez prekrivanja z obveznimi predmeti v isti skupini,
    - čim bližje originalnemu datumu.
    """
    predmet = row["Predmet"]
    rok = row["Rok"]
    orig_date = row["Datum"]

    if rok is None or orig_date is None:
        return []

    start, end = get_rok_range(rok)
    if start is None or end is None:
        return []

    # druge datume istega predmeta v ISTEM roku
    other_dates = [
        d
        for idx, d in context["dates_by_predmet_rok"].get((predmet, rok), [])
        if idx != row.name
    ]

    # vsi drugi izpiti v istem roku (mapa rok -> datum -> list zapisov)
    rok_to_dates = context["rows_by_rok_date"].get(rok, {})

    predmet_groups = predmet_groups_for_rok(rok, predmet)

    candidates = []

    current = start
    while current <= end:
        if current == orig_date:
            current += timedelta(days=1)
            continue

        if is_weekend(current):
            current += timedelta(days=1)
            continue
        if current in PRAZNIKI:
            current += timedelta(days=1)
            continue

        # 10-dnevno pravilo
        conflict_10 = False
        for d in other_dates:
            if abs((current - d).days) < 10:
                conflict_10 = True
                break
        if conflict_10:
            current += timedelta(days=1)
            continue

        # prekrivanje obveznih – samo, če je predmet v kakšni skupini
        if predmet_groups:
            df_same_day = rok_to_dates.get(current, [])
            conflict_obv = False
            for r2 in df_same_day:
                for grp in predmet_groups:
                    if r2["Predmet"] in grp:
                        conflict_obv = True
                        break
                if conflict_obv:
                    break
            if conflict_obv:
                current += timedelta(days=1)
                continue

        dist = abs((current - orig_date).days)
        candidates.append((current, dist))

        current += timedelta(days=1)

    if not candidates:
        return []

    candidates_sorted = sorted(candidates, key=lambda x: (x[1], x[0]))
    return [c[0] for c in candidates_sorted[:top_n]]


# =============================
# 1. BRANJE IN NORMALIZACIJA
# =============================

df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
df = df.copy()

# normalizacija imen predmetov v surovem df
df["Predmeti"] = df["Predmeti"].apply(normalize_predmet_name)

df["Datum_norm"] = df["Datum"].apply(normalize_date)
df["Rok"] = df["Datum_norm"].apply(get_rok)

# NAPAKE_PARSANJA
parsing_issues = []
for _, r in df.iterrows():
    orig_datum = r.get("Datum", None)
    norm_datum = r.get("Datum_norm", None)
    predmeti = r.get("Predmeti", "")
    prof = r.get("Prof.", "")

    issue = None
    if pd.notna(orig_datum) and norm_datum is None:
        issue = "Datum_neprepoznan"
    elif norm_datum is not None and get_rok(norm_datum) is None:
        issue = "Datum_brez_roka"

    if not predmeti or str(predmeti).strip() == "":
        issue = (issue + "; Prazen_predmet") if issue else "Prazen_predmet"

    if issue:
        parsing_issues.append({
            "Prof": prof,
            "Predmeti_raw": predmeti,
            "Datum_raw": orig_datum,
            "Datum_norm": norm_datum,
            "Opis": issue,
        })

df_parsing_issues = pd.DataFrame(parsing_issues)

expanded_rows = []
for _, row in df.iterrows():
    expanded = expand_predmeti_row(row)
    expanded_rows.extend(expanded)

if not expanded_rows:
    raise SystemExit("Ni podatkov za obdelavo – preveri stolpec 'Predmeti'.")

df_norm = pd.DataFrame(expanded_rows)

df_norm = df_norm[
    [
        "Prof.",
        "Predmet",
        "Datum_norm",
        "Rok",
        "Ura",
        "Predavalnica",
        "Dejanska predavalnica",
        "Opomba",
    ]
].rename(columns={"Prof.": "Prof", "Datum_norm": "Datum"})

df_norm = df_norm.reset_index(drop=True)

# filter po roku / profesorju, če podan v argumentih
if args.rok_filter:
    df_norm = df_norm[df_norm["Rok"] == args.rok_filter]

if args.prof_filter:
    df_norm = df_norm[df_norm["Prof"].astype(str).str.contains(args.prof_filter, case=False, na=False)]

# =============================
# 2. OSNOVNA PREVERJANJA
# =============================

df_norm["is_vikend"] = df_norm["Datum"].apply(is_weekend)
df_norm["is_izven"] = df_norm["Rok"].isna() & df_norm["Datum"].notna()
df_norm["is_praznik"] = df_norm["Datum"].isin(PRAZNIKI)

# 10-dnevno pravilo po predmetu in ROKU
rows_10dni = []
problem_10_pairs = set()  # (Predmet, Rok, Datum)
df_with_dates_10 = df_norm[df_norm["Datum"].notna() & df_norm["Rok"].notna()]

for (predmet, rok), grp in df_with_dates_10.groupby(["Predmet", "Rok"]):
    grp_sorted = grp.sort_values("Datum")
    dates = grp_sorted["Datum"].tolist()
    for i in range(len(dates) - 1):
        d1 = dates[i]
        d2 = dates[i + 1]
        delta = (d2 - d1).days
        if delta < 10:
            rows_10dni.append(
                {
                    "Predmet": predmet,
                    "Rok": rok,
                    "Prof_1": grp_sorted.iloc[i]["Prof"],
                    "Datum_1": d1,
                    "Prof_2": grp_sorted.iloc[i + 1]["Prof"],
                    "Datum_2": d2,
                    "Razlika_dni": delta,
                }
            )
            problem_10_pairs.add((predmet, rok, d1))
            problem_10_pairs.add((predmet, rok, d2))

df_10dni = pd.DataFrame(rows_10dni)
df_norm["is_10dni"] = df_norm.apply(
    lambda r: (r["Predmet"], r["Rok"], r["Datum"]) in problem_10_pairs,
    axis=1
)

# =============================
# 3. PREKRIVANJE OBVEZNIH PREDMETOV
# =============================

conflict_records = []
seen_pairs = set()
prekrivanje_keys = set()  # (Rok, Datum, Predmet, Prof)

for rok, skupine in KONFLIKTNE_SKUPINE_PO_ROKU.items():
    df_r = df_norm[(df_norm["Rok"] == rok) & df_norm["Datum"].notna()].copy()
    if df_r.empty:
        continue

    for datum, df_dan in df_r.groupby("Datum"):
        for idx_skupine, skupina in enumerate(skupine, start=1):
            df_conf = df_dan[df_dan["Predmet"].isin(skupina)].copy()
            if len(df_conf) < 2:
                continue

            rows = df_conf.reset_index(drop=True)
            for i in range(len(rows)):
                for j in range(i + 1, len(rows)):
                    a = rows.iloc[i]
                    b = rows.iloc[j]

                    key = (rok, datum, a["Predmet"], b["Predmet"], a["Prof"], b["Prof"])
                    key_rev = (rok, datum, b["Predmet"], a["Predmet"], b["Prof"], a["Prof"])

                    if key in seen_pairs or key_rev in seen_pairs:
                        continue
                    seen_pairs.add(key)

                    conflict_records.append(
                        {
                            "Rok": rok,
                            "Datum": datum,
                            "Skupina_id": f"{rok}_{idx_skupine}",
                            "Predmet_A": a["Predmet"],
                            "Prof_A": a["Prof"],
                            "Ura_A": a["Ura"],
                            "Predavalnica_A": a["Predavalnica"],
                            "Predmet_B": b["Predmet"],
                            "Prof_B": b["Prof"],
                            "Ura_B": b["Ura"],
                            "Predavalnica_B": b["Predavalnica"],
                        }
                    )
                    prekrivanje_keys.add((rok, datum, a["Predmet"], a["Prof"]))
                    prekrivanje_keys.add((rok, datum, b["Predmet"], b["Prof"]))

df_prekrivanja = pd.DataFrame(conflict_records)

df_norm["is_prekrivanje"] = df_norm.apply(
    lambda r: (r["Rok"], r["Datum"], r["Predmet"], r["Prof"]) in prekrivanje_keys,
    axis=1
)

# =============================
# 4. POMOŽNE STRUKTURE ZA PODROBNOSTI
# =============================

# 10-dni: map (Predmet, Rok, Datum) -> set(drugega datuma)
ten_dict = defaultdict(set)
for _, r in df_10dni.iterrows():
    key1 = (r["Predmet"], r["Rok"], r["Datum_1"])
    key2 = (r["Predmet"], r["Rok"], r["Datum_2"])
    ten_dict[key1].add(r["Datum_2"])
    ten_dict[key2].add(r["Datum_1"])

# prekrivanja: map (Rok, Datum, Predmet, Prof) -> set("Drug predmet (Prof)")
prekriv_dict = defaultdict(set)
for _, r in df_prekrivanja.iterrows():
    keyA = (r["Rok"], r["Datum"], r["Predmet_A"], r["Prof_A"])
    keyB = (r["Rok"], r["Datum"], r["Predmet_B"], r["Prof_B"])
    prekriv_dict[keyA].add(f"{r['Predmet_B']} ({r['Prof_B']})")
    prekriv_dict[keyB].add(f"{r['Predmet_A']} ({r['Prof_A']})")

# =============================
# 5. PREDLOGI NOVIH DATUMOV
# =============================

df_norm["is_problem"] = (
    df_norm["is_vikend"]
    | df_norm["is_izven"]
    | df_norm["is_praznik"]
    | df_norm["is_10dni"]
    | df_norm["is_prekrivanje"]
)

proposal_context = build_proposal_context(df_norm)
proposal_rows = []

for idx, row in df_norm[df_norm["is_problem"]].iterrows():
    reasons = []
    rok = row["Rok"]
    datum = row["Datum"]

    detail_10 = ""
    detail_prekriv = ""
    detail_izven = ""

    if row["is_vikend"]:
        reasons.append("vikend")
    if row["is_izven"]:
        reasons.append("izven_obdobja")
        if datum is not None:
            detail_izven = describe_out_of_range(datum)
    if row["is_praznik"]:
        reasons.append("praznik")
    if row["is_10dni"]:
        reasons.append("10_dni")
        key10 = (row["Predmet"], rok, datum)
        others = sorted(ten_dict.get(key10, []))
        if others:
            detail_10 = ", ".join(str(d) for d in others)
    if row["is_prekrivanje"]:
        reasons.append("prekrivanje_obveznih")
        keyp = (rok, datum, row["Predmet"], row["Prof"])
        others = sorted(prekriv_dict.get(keyp, []))
        if others:
            detail_prekriv = "; ".join(others)

    preds = propose_dates_for_row(row, proposal_context, top_n=5)

    rec = {
        "Prof": row["Prof"],
        "Predmet": row["Predmet"],
        "Stari_datum": row["Datum"],
        "Rok": rok,
        "Razlog": "; ".join(reasons) if reasons else "",
        "Podrobnost_10_dni": detail_10,
        "Podrobnost_prekrivanje": detail_prekriv,
        "Podrobnost_izven": detail_izven,
        "Predlog_1": preds[0] if len(preds) > 0 else None,
        "Predlog_2": preds[1] if len(preds) > 1 else None,
        "Predlog_3": preds[2] if len(preds) > 2 else None,
        "Predlog_4": preds[3] if len(preds) > 3 else None,
        "Predlog_5": preds[4] if len(preds) > 4 else None,
    }
    proposal_rows.append(rec)

df_predlogi = pd.DataFrame(proposal_rows)

if not df_predlogi.empty:
    df_predlogi = df_predlogi.drop_duplicates(
        subset=[
            "Prof",
            "Predmet",
            "Stari_datum",
            "Rok",
            "Razlog",
            "Podrobnost_10_dni",
            "Podrobnost_prekrivanje",
            "Podrobnost_izven",
        ]
    )
    df_predlogi = df_predlogi.sort_values(
        by=["Rok", "Stari_datum", "Prof", "Predmet"]
    ).reset_index(drop=True)

# set problematičnih za koledar
problem_keys = set(
    df_norm[df_norm["is_problem"] & df_norm["Datum"].notna()][["Rok", "Predmet", "Datum"]]
    .apply(lambda r: (r["Rok"], r["Predmet"], r["Datum"]), axis=1)
    .tolist()
)

# =============================
# 6. KOLEDAR OBVEZNIH + VSEH
# =============================

def generiraj_koledar(df_norm: pd.DataFrame, problem_keys: set, out_name: str | None = None):
    """
    Iz df_norm generira:
    - koledar obveznih predmetov po rokih (barve po konfliktnih skupinah,
      dodatna barva v stolpcu B, datumi od stolpca C naprej),
    - koledar vseh predmetov po rokih (vsi termini svetlo modro),
    - LEGENDA list.
    """
    # močno kontrastne barve za skupine (ARGB)
    COLORS = [
        "00FF9999",
        "0099CCFF",
        "00FFFF99",
        "00FFCC00",
        "0099FF99",
        "00CC99FF",
        "00FF99CC",
        "0000CCFF",
        "00FF6600",
        "0066FF66",
        "00CC0066",
        "006666FF",
        "00CCFF00",
        "00FF00CC",
        "00CC3300",
        "003399CC",
    ]
    LIGHT_BLUE = PatternFill(start_color="00CFE2F3", end_color="00CFE2F3", fill_type="solid")
    BROWN_FILL = PatternFill(start_color="00865C3C", end_color="00865C3C", fill_type="solid")

    # obvezni predmeti + mapping predmet -> skupine
    obvezni_predmeti = {}
    predmet_groups = {}
    group_colors = {}
    all_obvezni = set()

    for rok, skupine in KONFLIKTNE_SKUPINE_PO_ROKU.items():
        s = set()
        pmap = defaultdict(list)
        for idx, grp in enumerate(skupine):
            for predmet in grp:
                s.add(predmet)
                pmap[predmet].append(idx)
        obvezni_predmeti[rok] = s
        predmet_groups[rok] = pmap

        for idx, grp in enumerate(skupine):
            color = COLORS[idx % len(COLORS)]
            group_colors[(rok, idx)] = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

        all_obvezni.update(s)

    # za jesensko – vzemi vse obvezne, a brez posebnih skupin
    if "jesensko" not in obvezni_predmeti:
        obvezni_predmeti["jesensko"] = all_obvezni
        predmet_groups["jesensko"] = defaultdict(list)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    red_border = Border(
        left=Side(style="medium", color="00FF0000"),
        right=Side(style="medium", color="00FF0000"),
        top=Side(style="medium", color="00FF0000"),
        bottom=Side(style="medium", color="00FF0000"),
    )

    def format_sheet(ws, filter_only_col_A=False):
        max_row = ws.max_row
        if filter_only_col_A:
            ws.auto_filter.ref = f"A1:A{max_row}"
        else:
            ws.auto_filter.ref = ws.dimensions

        ws.freeze_panes = "A2"

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            col = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    value = str(cell.value) if cell.value is not None else ""
                except Exception:
                    value = ""
                if len(value) > max_length:
                    max_length = len(value)
            ws.column_dimensions[col].width = max_length + 2

    # -------- koledar OBVEZNIH --------
    for rok, sheetname in [("zimsko", "zimsko"), ("poletno", "poletno"), ("jesensko", "jesensko")]:
        start, end = get_rok_range(rok)
        if start is None or end is None:
            continue

        ws = wb.create_sheet(title=sheetname)

        ws.cell(row=1, column=1).value = "Predmet"
        # B1 pustimo za dodatno barvo
        dates = []
        cur = start
        col_idx = 3  # datumi od C naprej
        while cur <= end:
            cell = ws.cell(row=1, column=col_idx)
            cell.value = cur
            dates.append(cur)

            # vikendi + prazniki rjavo
            if is_weekend(cur) or cur in PRAZNIKI:
                cell.fill = BROWN_FILL

            col_idx += 1
            cur += timedelta(days=1)

        date_to_col = {d: i + 3 for i, d in enumerate(dates)}

        obv_set = obvezni_predmeti.get(rok, set())
        df_r = df_norm[
            (df_norm["Rok"] == rok)
            & (df_norm["Predmet"].isin(obv_set))
            & df_norm["Datum"].notna()
        ].copy()

        if df_r.empty:
            format_sheet(ws, filter_only_col_A=True)
            continue

        predmeti_v_roku = sorted(df_r["Predmet"].unique())
        row_idx = 2

        for predmet in predmeti_v_roku:
            ws.cell(row=row_idx, column=1).value = predmet

            groups = predmet_groups.get(rok, {}).get(predmet, [])
            main_fill = None
            extra_fill = None

            if groups:
                main_fill = group_colors.get((rok, groups[0]))
                if len(groups) > 1:
                    extra_fill = group_colors.get((rok, groups[1]))

            if main_fill is not None:
                ws.cell(row=row_idx, column=1).fill = main_fill
            if extra_fill is not None:
                ws.cell(row=row_idx, column=2).fill = extra_fill

            df_p = df_r[df_r["Predmet"] == predmet]
            for _, rec in df_p.iterrows():
                d = rec["Datum"]
                col = date_to_col.get(d)
                if col is None:
                    continue
                cell = ws.cell(row=row_idx, column=col)
                cell.value = predmet
                if main_fill is not None:
                    cell.fill = main_fill
                if (rok, predmet, d) in problem_keys:
                    cell.border = red_border

            row_idx += 1

        # opcijsko obarvaj vikend/praznik stolpce po celotnem stolpcu
        max_row = ws.max_row
        for d, col in date_to_col.items():
            if is_weekend(d) or d in PRAZNIKI:
                for r in range(2, max_row + 1):
                    cell = ws.cell(row=r, column=col)
                    if cell.fill is None or cell.fill.fill_type is None:
                        cell.fill = BROWN_FILL

        format_sheet(ws, filter_only_col_A=True)

    # -------- koledar VSEH predmetov --------
    for rok, sheetname in [("zimsko", "zimsko_vsi"), ("poletno", "poletno_vsi"), ("jesensko", "jesensko_vsi")]:
        start, end = get_rok_range(rok)
        if start is None or end is None:
            continue

        ws = wb.create_sheet(title=sheetname)

        ws.cell(row=1, column=1).value = "Predmet"
        dates = []
        cur = start
        col_idx = 2
        while cur <= end:
            cell = ws.cell(row=1, column=col_idx)
            cell.value = cur
            dates.append(cur)

            if is_weekend(cur) or cur in PRAZNIKI:
                cell.fill = BROWN_FILL

            col_idx += 1
            cur += timedelta(days=1)

        date_to_col = {d: i + 2 for i, d in enumerate(dates)}

        df_r = df_norm[
            (df_norm["Rok"] == rok)
            & df_norm["Datum"].notna()
        ].copy()

        if df_r.empty:
            format_sheet(ws, filter_only_col_A=True)
            continue

        predmeti_v_roku = sorted(df_r["Predmet"].unique())
        row_idx = 2

        for predmet in predmeti_v_roku:
            ws.cell(row=row_idx, column=1).value = predmet
            df_p = df_r[df_r["Predmet"] == predmet]
            for _, rec in df_p.iterrows():
                d = rec["Datum"]
                col = date_to_col.get(d)
                if col is None:
                    continue
                cell = ws.cell(row=row_idx, column=col)
                cell.value = predmet
                cell.fill = LIGHT_BLUE
                if (rok, predmet, d) in problem_keys:
                    cell.border = red_border
            row_idx += 1

        max_row = ws.max_row
        for d, col in date_to_col.items():
            if is_weekend(d) or d in PRAZNIKI:
                for r in range(2, max_row + 1):
                    cell = ws.cell(row=r, column=col)
                    if cell.fill is None or cell.fill.fill_type is None:
                        cell.fill = BROWN_FILL

        format_sheet(ws, filter_only_col_A=True)

    # -------- LEGENDA --------
    ws_leg = wb.create_sheet(title="LEGENDA")
    ws_leg["A1"] = "Rok"
    ws_leg["B1"] = "ID_skupine"
    ws_leg["C1"] = "Predmeti v skupini"
    ws_leg["D1"] = "Barva (ARGB)"

    row = 2
    for rok, skupine in KONFLIKTNE_SKUPINE_PO_ROKU.items():
        for idx, grp in enumerate(skupine):
            fill = group_colors.get((rok, idx))
            color_hex = None
            if fill is not None and fill.start_color is not None:
                color_hex = fill.start_color.rgb
            ws_leg.cell(row=row, column=1).value = rok
            ws_leg.cell(row=row, column=2).value = idx
            ws_leg.cell(row=row, column=3).value = ", ".join(grp)
            ws_leg.cell(row=row, column=4).value = color_hex
            if fill is not None:
                ws_leg.cell(row=row, column=3).fill = fill
            row += 1

    format_sheet(ws_leg, filter_only_col_A=False)

    # zapis datoteke
    if out_name is not None:
        output_file = os.path.join(REPORT_DIR, out_name)
    else:
        base_prefix = "izpitni_koledar_OBVEZNI_auto"
        output_file = get_next_version_path(base_prefix)

    # prestavi stare koledarje v arhiv
    archive_old_files("izpitni_koledar_OBVEZNI_auto", output_file)

    wb.save(output_file)
    print(f"Izpitni koledar zapisan v: {output_file}")
    try_open_file(output_file, "Koledarja")

    return output_file




# =============================
# 7. ZAPIS POROČILA
# =============================

# ---- IZRAČUN IMENA POROČILA ----
if args.out_porocilo is not None:
    output_file = os.path.join(REPORT_DIR, args.out_porocilo)
else:
    base_prefix = "porocilo_izpiti"
    output_file = get_next_version_path(base_prefix)

# ---- ARHIVIRAJ STARA POROČILA ----
archive_old_files("porocilo_izpiti", output_file)


with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    workbook = writer.book

    # 1) PREDLOGI_DATUMOV – prvi zavihek
    if not df_predlogi.empty:
        df_predlogi.to_excel(writer, sheet_name="PREDLOGI_DATUMOV", index=False)

    # 2) osnovna normalizirana tabela
    df_norm.to_excel(writer, sheet_name="IZPITI_NORMALIZIRANO", index=False)

    # 3) napake
    df_vikend = df_norm[df_norm["is_vikend"]].copy()
    df_izven = df_norm[df_norm["is_izven"]].copy()
    df_prazniki = df_norm[df_norm["is_praznik"]].copy()

    if not df_vikend.empty:
        df_vikend.to_excel(writer, sheet_name="NAPAKE_VIKEND", index=False)
    if not df_izven.empty:
        df_izven.to_excel(writer, sheet_name="NAPAKE_IZVEN_OBDOBJA", index=False)
    if not df_prazniki.empty:
        df_prazniki.to_excel(writer, sheet_name="NAPAKE_PRAZNIK", index=False)
    if not df_10dni.empty:
        df_10dni.to_excel(writer, sheet_name="NAPAKE_10_DNI", index=False)
    if not df_prekrivanja.empty:
        df_prekrivanja.to_excel(writer, sheet_name="NAPAKE_PREKRIVANJA", index=False)
    if not df_parsing_issues.empty:
        df_parsing_issues.to_excel(writer, sheet_name="NAPAKE_PARSANJA", index=False)

    # ===== barvanje skupin prekrivanj v PREDLOGI_DATUMOV =====
    if "PREDLOGI_DATUMOV" in workbook.sheetnames:
        ws = workbook["PREDLOGI_DATUMOV"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=1, column=col).value] = col

        rok_col = headers.get("Rok")
        date_col = headers.get("Stari_datum")
        razlog_col = headers.get("Razlog")

        if rok_col and date_col and razlog_col:
            groups = defaultdict(list)
            for row_i in range(2, ws.max_row + 1):
                razlog_val = ws.cell(row=row_i, column=razlog_col).value
                if not razlog_val or "prekrivanje_obveznih" not in str(razlog_val):
                    continue
                rok_val = ws.cell(row=row_i, column=rok_col).value
                dat_val = ws.cell(row=row_i, column=date_col).value
                key = (rok_val, dat_val)
                groups[key].append(row_i)

            fills = [
                "FFF2CC",
                "CCE5FF",
                "E2EFDA",
                "FCE4D6",
                "E4DFEC",
            ]
            color_idx = 0

            for key, rows in groups.items():
                if len(rows) < 2:
                    continue
                color = fills[color_idx % len(fills)]
                color_idx += 1
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for r in rows:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill

    # ===== auto filter + wrap + width za vse sheete =====
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            col = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    value = str(cell.value) if cell.value is not None else ""
                except Exception:
                    value = ""
                if len(value) > max_length:
                    max_length = len(value)
            ws.column_dimensions[col].width = max_length + 2

print(f"Poročilo zapisano v: {output_file}")


# poskusi odpreti poročilo
try_open_file(output_file, "Poročila")

# =============================
# 8. GENERIRAJ KOLEDAR
# =============================

koledar_name = args.out_koledar if args.out_koledar is not None else None
generiraj_koledar(df_norm, problem_keys, out_name=koledar_name)
