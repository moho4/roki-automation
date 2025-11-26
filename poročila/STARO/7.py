import os
import pandas as pd
from datetime import datetime, date, timedelta
from collections import defaultdict
from openpyxl.styles import PatternFill, Alignment

# =============================
# KONFIGURACIJA
# =============================

INPUT_FILE = "izpitni_roki_25_26_prilagojen_v5.xlsx"  # ime vhodne datoteke
SHEET_NAME = "Profesorji_pripravljeno"

# izpitna obdobja 2025/26
ZIMSKO_OD = date(2026, 1, 19)
ZIMSKO_DO = date(2026, 2, 13)

POLETNO_OD = date(2026, 6, 1)   # tvoj "spomladanski rok"
POLETNO_DO = date(2026, 7, 3)

JESENSKO_OD = date(2026, 8, 19)
JESENSKO_DO = date(2026, 9, 15)

# prazniki, kjer izpitov nočeš
PRAZNIKI = {
    date(2026, 2, 8),   # Prešernov dan
    date(2026, 6, 25),  # Dan državnosti
}

# =============================
# KONFLIKTNE SKUPINE – OBVEZNI
# =============================

# 1. stopnja – zimsko
KONFLIKTNE_SKUPINE_ZIMSKO = [
    # 1. letnik
    ["Antična filozofija 1", "Logika in argumentacija", "Novoveška filozofija 1", "Politična filozofija"],

    # 2. letnik E
    ["Srednjeveška in renesančna filozofija 1", "Filozofska antropologija", "Etika",
     "Azijske filozofije", "Slovenska filozofija in filozofska terminologija",
     "Marksizem in kritična teorija", "Filozofija in zgodovina znanosti"],

    # 2. letnik D – Filozofija splošno
    ["Srednjeveška in renesančna filozofija 1", "Filozofska antropologija",
     "Strukturalizem, psihoanaliza, filozofija"],

    # 2. letnik D – Kultura in etika
    ["Etika", "Moralna filozofija"],

    # 3. letnik E
    ["Fenomenologija 1", "Praktična filozofija", "Nemška klasična filozofija E",
     "Hermenevtika 2", "Filozofija narave", "Sodobna analitična filozofija", "Metafizika"],

    # 3. letnik D – Filozofija splošno
    ["Fenomenologija 1", "Nemška klasična filozofija D", "Filozofija zavesti in življenja"],

    # 3. letnik D – Kultura in etika
    ["Praktična filozofija", "Filozofija in humanistika", "Filozofija religije"],

    # 2. stopnja pedagoška – zimsko
    ["Kritična teorija družbe", "Didaktika filozofskih praks"],
    ["Praktična etika", "Didaktika filozofije in etike"],
]

# 1. stopnja – poletno (spomladanski rok)
KONFLIKTNE_SKUPINE_POLETNO = [
    # 1. letnik
    ["Estetika", "Antična filozofija 2", "Simbolna logika",
     "Spoznavna teorija", "Novoveška filozofija 2", "Ontologija"],

    # 2. letnik E
    ["Hermenevtika 1", "Osnove analitične filozofije", "Srednjeveška in renesančna filozofija 2",
     "Uvod v psihoanalizo", "Filozofija duha"],

    # 2. letnik D – Filozofija splošno
    ["Hermenevtika 1", "Osnove analitične filozofije", "Filozofija zgodovine"],

    # 2. letnik D – Kultura in etika
    ["Estetika", "Azijske filozofije, religije in kulture", "Socialna filozofija"],

    # 3. letnik E
    ["Normativna etika in teorija delovanja", "Fenomenologija 2",
     "Filozofija jezika", "Antropologija simbolnih form"],

    # 3. letnik D – Filozofija splošno
    ["Normativna etika in teorija delovanja", "Semiotika"],

    # 3. letnik D – Kultura in etika
    ["Praktična filozofija med Kantom in Heglom", "Človek in kozmos v renesansi"],
]

KONFLIKTNE_SKUPINE_PO_ROKU = {
    "zimsko": KONFLIKTNE_SKUPINE_ZIMSKO,
    "poletno": KONFLIKTNE_SKUPINE_POLETNO,
    # "jesensko": []  # trenutno brez striktnih skupin
}

# =============================
# POMOŽNE FUNKCIJE
# =============================

def get_rok(d: date) -> str | None:
    if d is None:
        return None
    if ZIMSKO_OD <= d <= ZIMSKO_DO:
        return "zimsko"
    if POLETNO_OD <= d <= POLETNO_DO:
        return "poletno"
    if JESENSKO_OD <= d <= JESENSKO_DO:
        return "jesensko"
    return None


def get_rok_range(rok: str):
    if rok == "zimsko":
        return ZIMSKO_OD, ZIMSKO_DO
    if rok == "poletno":
        return POLETNO_OD, POLETNO_DO
    if rok == "jesensko":
        return JESENSKO_OD, JESENSKO_DO
    return None, None


def describe_out_of_range(d: date) -> str:
    """Opis, katero mejo roka datum krši."""
    if d < ZIMSKO_OD:
        return f"pred ZIMSKO_OD ({ZIMSKO_OD})"
    if ZIMSKO_DO < d < POLETNO_OD:
        return f"po ZIMSKO_DO ({ZIMSKO_DO})"
    if POLETNO_DO < d < JESENSKO_OD:
        return f"po POLETNO_DO ({POLETNO_DO})"
    if d > JESENSKO_DO:
        return f"po JESENSKO_DO ({JESENSKO_DO})"
    return "znotraj katerega od obdobij (logična napaka)"


def normalize_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%d. %m. %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def expand_predmeti_row(row):
    """
    'Predmeti' = 'A. B. C' -> tri vrstice: A, B, C (isti datum, prof, ura ...).
    """
    predmeti_raw = row["Predmeti"]
    if pd.isna(predmeti_raw):
        return []

    parts = [p.strip() for p in str(predmeti_raw).split(".") if p.strip()]
    expanded = []
    for predmet in parts:
        new_row = row.copy()
        new_row["Predmet"] = predmet
        expanded.append(new_row)
    return expanded


def is_weekend(d: date) -> bool:
    if d is None:
        return False
    return d.weekday() >= 5  # 5 = sobota, 6 = nedelja


# =============================
# 1. BRANJE IN NORMALIZACIJA
# =============================

df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
df = df.copy()

df["Datum_norm"] = df["Datum"].apply(normalize_date)
df["Rok"] = df["Datum_norm"].apply(get_rok)

expanded_rows = []
for _, row in df.iterrows():
    expanded = expand_predmeti_row(row)
    expanded_rows.extend(expanded)

if not expanded_rows:
    raise SystemExit("Ni podatkov za obdelavo – preveri stolpec 'Predmeti'.")

df_norm = pd.DataFrame(expanded_rows)

df_norm = df_norm[
    [
        "Prof.",
        "Predmet",
        "Datum_norm",
        "Rok",
        "Ura",
        "Predavalnica",
        "Dejanska predavalnica",
        "Opomba",
    ]
].rename(columns={"Prof.": "Prof", "Datum_norm": "Datum"})

df_norm = df_norm.reset_index(drop=True)

# =============================
# 2. OSNOVNA PREVERJANJA
# =============================

df_norm["is_vikend"] = df_norm["Datum"].apply(is_weekend)
df_norm["is_izven"] = df_norm["Rok"].isna() & df_norm["Datum"].notna()
df_norm["is_praznik"] = df_norm["Datum"].isin(PRAZNIKI)

# 10-dnevno pravilo po predmetu in ROKU (znotraj istega obdobja)
rows_10dni = []
problem_10_pairs = set()  # (Predmet, Rok, Datum)
df_with_dates_10 = df_norm[df_norm["Datum"].notna() & df_norm["Rok"].notna()]

for (predmet, rok), grp in df_with_dates_10.groupby(["Predmet", "Rok"]):
    grp_sorted = grp.sort_values("Datum")
    dates = grp_sorted["Datum"].tolist()
    for i in range(len(dates) - 1):
        d1 = dates[i]
        d2 = dates[i + 1]
        delta = (d2 - d1).days
        if delta < 10:
            rows_10dni.append(
                {
                    "Predmet": predmet,
                    "Rok": rok,
                    "Prof_1": grp_sorted.iloc[i]["Prof"],
                    "Datum_1": d1,
                    "Prof_2": grp_sorted.iloc[i + 1]["Prof"],
                    "Datum_2": d2,
                    "Razlika_dni": delta,
                }
            )
            problem_10_pairs.add((predmet, rok, d1))
            problem_10_pairs.add((predmet, rok, d2))

df_10dni = pd.DataFrame(rows_10dni)
df_norm["is_10dni"] = df_norm.apply(
    lambda r: (r["Predmet"], r["Rok"], r["Datum"]) in problem_10_pairs,
    axis=1
)

# =============================
# 3. PREKRIVANJE OBVEZNIH PREDMETOV
# =============================

conflict_records = []
seen_pairs = set()
prekrivanje_keys = set()  # (Rok, Datum, Predmet, Prof)

for rok, skupine in KONFLIKTNE_SKUPINE_PO_ROKU.items():
    df_r = df_norm[(df_norm["Rok"] == rok) & df_norm["Datum"].notna()].copy()
    if df_r.empty:
        continue

    for datum, df_dan in df_r.groupby("Datum"):
        for idx_skupine, skupina in enumerate(skupine, start=1):
            df_conf = df_dan[df_dan["Predmet"].isin(skupina)].copy()
            if len(df_conf) < 2:
                continue

            rows = df_conf.reset_index(drop=True)
            for i in range(len(rows)):
                for j in range(i + 1, len(rows)):
                    a = rows.iloc[i]
                    b = rows.iloc[j]

                    key = (rok, datum, a["Predmet"], b["Predmet"], a["Prof"], b["Prof"])
                    key_rev = (rok, datum, b["Predmet"], a["Predmet"], b["Prof"], a["Prof"])

                    if key in seen_pairs or key_rev in seen_pairs:
                        continue
                    seen_pairs.add(key)

                    conflict_records.append(
                        {
                            "Rok": rok,
                            "Datum": datum,
                            "Skupina_id": f"{rok}_{idx_skupine}",
                            "Predmet_A": a["Predmet"],
                            "Prof_A": a["Prof"],
                            "Ura_A": a["Ura"],
                            "Predavalnica_A": a["Predavalnica"],
                            "Predmet_B": b["Predmet"],
                            "Prof_B": b["Prof"],
                            "Ura_B": b["Ura"],
                            "Predavalnica_B": b["Predavalnica"],
                        }
                    )
                    prekrivanje_keys.add((rok, datum, a["Predmet"], a["Prof"]))
                    prekrivanje_keys.add((rok, datum, b["Predmet"], b["Prof"]))

df_prekrivanja = pd.DataFrame(conflict_records)

df_norm["is_prekrivanje"] = df_norm.apply(
    lambda r: (r["Rok"], r["Datum"], r["Predmet"], r["Prof"]) in prekrivanje_keys,
    axis=1
)

# =============================
# 4. POMOŽNE STRUKTURE ZA PODROBNOSTI
# =============================

# 10-dni: map (Predmet, Rok, Datum) -> set(drugega datuma)
ten_dict = defaultdict(set)
for _, r in df_10dni.iterrows():
    key1 = (r["Predmet"], r["Rok"], r["Datum_1"])
    key2 = (r["Predmet"], r["Rok"], r["Datum_2"])
    ten_dict[key1].add(r["Datum_2"])
    ten_dict[key2].add(r["Datum_1"])

# prekrivanja: map (Rok, Datum, Predmet, Prof) -> set("Drug predmet (Prof)")
prekriv_dict = defaultdict(set)
for _, r in df_prekrivanja.iterrows():
    keyA = (r["Rok"], r["Datum"], r["Predmet_A"], r["Prof_A"])
    keyB = (r["Rok"], r["Datum"], r["Predmet_B"], r["Prof_B"])
    prekriv_dict[keyA].add(f"{r['Predmet_B']} ({r['Prof_B']})")
    prekriv_dict[keyB].add(f"{r['Predmet_A']} ({r['Prof_A']})")

# =============================
# 5. PREDLOGI NOVIH DATUMOV (TOP 5, isti rok)
# =============================

def predmet_groups_for_rok(rok: str, predmet: str):
    """Vrne seznam skupin (list[str]) v katerih je ta predmet za dani rok."""
    if rok not in KONFLIKTNE_SKUPINE_PO_ROKU:
        return []
    groups = []
    for grp in KONFLIKTNE_SKUPINE_PO_ROKU[rok]:
        if predmet in grp:
            groups.append(grp)
    return groups


def propose_dates_for_row(row, df_all: pd.DataFrame, top_n: int = 5):
    """
    Za eno problematično vrstico predlaga do top_n novih datumov.
    Upošteva:
    - isti rok,
    - brez vikendov,
    - brez praznikov,
    - 10-dnevno pravilo znotraj istega roka (naprej + nazaj),
    - brez prekrivanja z obveznimi predmeti v isti skupini,
    - čim bližje originalnemu datumu.
    """
    predmet = row["Predmet"]
    rok = row["Rok"]
    orig_date = row["Datum"]

    if rok is None or orig_date is None:
        return []

    start, end = get_rok_range(rok)
    if start is None or end is None:
        return []

    # druge datume istega predmeta v ISTEM roku
    mask_same_predmet = (
        (df_all["Predmet"] == predmet) &
        df_all["Datum"].notna() &
        (df_all["Rok"] == rok)
    )
    other_dates = df_all[mask_same_predmet & (df_all.index != row.name)]["Datum"].tolist()

    # vsi drugi izpiti v istem roku
    df_same_rok = df_all[(df_all["Rok"] == rok) & df_all["Datum"].notna()].copy()

    predmet_groups = predmet_groups_for_rok(rok, predmet)

    candidates = []

    current = start
    while current <= end:
        # originalnega datuma ne ponujamo
        if current == orig_date:
            current += timedelta(days=1)
            continue

        # vikend / praznik
        if is_weekend(current):
            current += timedelta(days=1)
            continue
        if current in PRAZNIKI:
            current += timedelta(days=1)
            continue

        # 10-dnevno pravilo – samo znotraj tega roka
        conflict_10 = False
        for d in other_dates:
            if abs((current - d).days) < 10:
                conflict_10 = True
                break
        if conflict_10:
            current += timedelta(days=1)
            continue

        # prekrivanje obveznih – samo, če je predmet v kakšni skupini
        if predmet_groups:
            df_same_day = df_same_rok[df_same_rok["Datum"] == current]
            conflict_obv = False
            for _, r2 in df_same_day.iterrows():
                for grp in predmet_groups:
                    if r2["Predmet"] in grp:
                        conflict_obv = True
                        break
                if conflict_obv:
                    break
            if conflict_obv:
                current += timedelta(days=1)
                continue

        # kandidat OK
        dist = abs((current - orig_date).days)
        candidates.append((current, dist))

        current += timedelta(days=1)

    if not candidates:
        return []

    candidates_sorted = sorted(candidates, key=lambda x: (x[1], x[0]))
    return [c[0] for c in candidates_sorted[:top_n]]


df_norm["is_problem"] = (
    df_norm["is_vikend"] |
    df_norm["is_izven"] |
    df_norm["is_praznik"] |
    df_norm["is_10dni"] |
    df_norm["is_prekrivanje"]
)

proposal_rows = []

for idx, row in df_norm[df_norm["is_problem"]].iterrows():
    reasons = []
    rok = row["Rok"]
    datum = row["Datum"]

    detail_10 = ""
    detail_prekriv = ""
    detail_izven = ""

    if row["is_vikend"]:
        reasons.append("vikend")
    if row["is_izven"]:
        reasons.append("izven_obdobja")
        if datum is not None:
            detail_izven = describe_out_of_range(datum)
    if row["is_praznik"]:
        reasons.append("praznik")
    if row["is_10dni"]:
        reasons.append("10_dni")
        key10 = (row["Predmet"], rok, datum)
        others = sorted(ten_dict.get(key10, []))
        if others:
            detail_10 = ", ".join(str(d) for d in others)
    if row["is_prekrivanje"]:
        reasons.append("prekrivanje_obveznih")
        keyp = (rok, datum, row["Predmet"], row["Prof"])
        others = sorted(prekriv_dict.get(keyp, []))
        if others:
            detail_prekriv = "; ".join(others)

    preds = propose_dates_for_row(row, df_norm, top_n=5)

    rec = {
        "Prof": row["Prof"],
        "Predmet": row["Predmet"],
        "Stari_datum": row["Datum"],
        "Rok": rok,
        "Razlog": "; ".join(reasons) if reasons else "",
        "Podrobnost_10_dni": detail_10,
        "Podrobnost_prekrivanje": detail_prekriv,
        "Podrobnost_izven": detail_izven,
        "Predlog_1": preds[0] if len(preds) > 0 else None,
        "Predlog_2": preds[1] if len(preds) > 1 else None,
        "Predlog_3": preds[2] if len(preds) > 2 else None,
        "Predlog_4": preds[3] if len(preds) > 3 else None,
        "Predlog_5": preds[4] if len(preds) > 4 else None,
    }
    proposal_rows.append(rec)

df_predlogi = pd.DataFrame(proposal_rows)

# odstranimo duplikate in uredimo
if not df_predlogi.empty:
    df_predlogi = df_predlogi.drop_duplicates(
        subset=["Prof", "Predmet", "Stari_datum", "Rok", "Razlog",
                "Podrobnost_10_dni", "Podrobnost_prekrivanje", "Podrobnost_izven"]
    )
    df_predlogi = df_predlogi.sort_values(
        by=["Rok", "Stari_datum", "Prof", "Predmet"]
    ).reset_index(drop=True)

# =============================
# 6. ZAPIS POROČILA Z DINAMIČNIM IMENOM
# =============================

base_name = "porocilo_izpiti"
ext = ".xlsx"
output_file = base_name + ext

counter = 2
while os.path.exists(output_file):
    output_file = f"{base_name}_v{counter}{ext}"
    counter += 1

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    # osnovna normalizirana tabela
    df_norm.to_excel(writer, sheet_name="IZPITI_NORMALIZIRANO", index=False)

    # napake
    df_vikend = df_norm[df_norm["is_vikend"]].copy()
    df_izven = df_norm[df_norm["is_izven"]].copy()
    df_prazniki = df_norm[df_norm["is_praznik"]].copy()

    if not df_vikend.empty:
        df_vikend.to_excel(writer, sheet_name="NAPAKE_VIKEND", index=False)
    if not df_izven.empty:
        df_izven.to_excel(writer, sheet_name="NAPAKE_IZVEN_OBDOBJA", index=False)
    if not df_prazniki.empty:
        df_prazniki.to_excel(writer, sheet_name="NAPAKE_PRAZNIK", index=False)
    if not df_10dni.empty:
        df_10dni.to_excel(writer, sheet_name="NAPAKE_10_DNI", index=False)
    if not df_prekrivanja.empty:
        df_prekrivanja.to_excel(writer, sheet_name="NAPAKE_PREKRIVANJA", index=False)

    # predlogi datumov
    if not df_predlogi.empty:
        df_predlogi.to_excel(writer, sheet_name="PREDLOGI_DATUMOV", index=False)

    workbook = writer.book

    # =============================
    # BARVANJE SKUPIN PREKRIVANJ V PREDLOGIH
    # =============================
    if "PREDLOGI_DATUMOV" in workbook.sheetnames:
        ws = workbook["PREDLOGI_DATUMOV"]

        # map header -> column index
        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=1, column=col).value] = col

        rok_col = headers.get("Rok")
        date_col = headers.get("Stari_datum")
        razlog_col = headers.get("Razlog")

        if rok_col and date_col and razlog_col:
            groups = defaultdict(list)
            for row_i in range(2, ws.max_row + 1):
                razlog_val = ws.cell(row=row_i, column=razlog_col).value
                if not razlog_val or "prekrivanje_obveznih" not in str(razlog_val):
                    continue
                rok_val = ws.cell(row=row_i, column=rok_col).value
                dat_val = ws.cell(row=row_i, column=date_col).value
                key = (rok_val, dat_val)
                groups[key].append(row_i)

            fills = [
                "FFF2CC",  # light orange
                "CCE5FF",  # light blue
                "E2EFDA",  # light green
                "FCE4D6",  # light red
                "E4DFEC",  # light purple
            ]
            color_idx = 0

            for key, rows in groups.items():
                if len(rows) < 2:
                    continue
                color = fills[color_idx % len(fills)]
                color_idx += 1
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for r in rows:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill

# =============================
# AUTO FILTER + WRAP TEXT + AUTO WIDTH ZA VSE SHEETE
# =============================
for sheet_name in workbook.sheetnames:
    ws = workbook[sheet_name]

    # auto filter čez celoten range (glava + podatki)
    ws.auto_filter.ref = ws.dimensions

    # zamrznemo prvo vrstico
    ws.freeze_panes = "A2"

    # wrap text za vse celice z vsebino
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(wrap_text=True)

    # auto width stolpcev
    for column_cells in ws.columns:
        max_length = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            try:
                value = str(cell.value) if cell.value is not None else ""
            except Exception:
                value = ""
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col].width = max_length + 2

print(f"Poročilo zapisano v: {output_file}")

# =============================
# 7. SAMODEJNO ODPRI DATOTEKO
# =============================

try:
    os.startfile(output_file)   # Windows
except Exception as e:
    print("Datoteke ni bilo mogoče samodejno odpreti:", e)
